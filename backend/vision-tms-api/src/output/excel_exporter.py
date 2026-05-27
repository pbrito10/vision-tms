from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

from src.output.metrics_snapshot import MetricsSnapshot
from src.output.output_interface import OutputInterface
from src.output.session_output import output_stamp
from src.tracking.cycle_result import CycleResult
from src.tracking.order_matching import RESULT_IN_ORDER, OrderDiagnosis, diagnose_order
from src.tracking.task_event import TaskEvent

_BOTTLENECK_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_TASK_COLUMN = "Tarefa/Zona"
_CYCLE_COLUMNS = [
    "Nº Ciclo",
    "Inicio",
    "Fim",
    "Duracao (s)",
    "Resultado do sistema",
    "Sequencia registada",
    "Problema detetado",
    "Classificacao manual",
    "Observacoes",
]

_FORCED_LABEL: dict[bool, str] = {True: "Sim", False: "Nao"}


@dataclass(frozen=True)
class _EventExportRow:
    event: TaskEvent
    counts_as_interruption: bool = False


def _cycle_diagnosis(cycle_result: CycleResult):
    if cycle_result.sequence_in_order:
        return OrderDiagnosis(RESULT_IN_ORDER, "Sem problema detetado.")

    return diagnose_order(cycle_result.actual_sequence, cycle_result.expected_sequence)


def _format_sequence(sequence) -> str:
    cleaned_steps = []
    for step in sequence:
        step_name = str(step).strip()
        while step_name.startswith("->"):
            step_name = step_name[2:].strip()
        if step_name:
            cleaned_steps.append(step_name)
    return ", ".join(cleaned_steps)


class ExcelExporter(OutputInterface):
    """Exporta os dados da sessao para um ficheiro .xlsx no fim."""

    def __init__(self, output_dir: Path, session_start: datetime) -> None:
        self._output_dir = output_dir
        self._session_start = session_start
        self._events: list[_EventExportRow] = []
        self._cycle_results: dict[int, CycleResult] = {}

        output_dir.mkdir(parents=True, exist_ok=True)

    def add_event(self, event: TaskEvent, counts_as_interruption: bool = False) -> None:
        """Acumula TaskEvents durante a sessao para exportar no fim."""
        self._events.append(_EventExportRow(event, counts_as_interruption))

    def add_cycle_result(self, cycle_result: CycleResult) -> None:
        """Regista o ciclo fechado para exportacao."""
        self._cycle_results[cycle_result.cycle_number] = cycle_result

    def write(self, snapshot: MetricsSnapshot) -> None:
        """Gera o ficheiro Excel com todas as folhas."""
        filename = f"sessao_{output_stamp(self._output_dir, self._session_start)}.xlsx"
        path = self._output_dir / filename

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            self._write_summary(writer, snapshot)
            self._write_zone_metrics(writer, snapshot)
            self._write_cycles(writer)
            self._write_events(writer)

    def _write_summary(self, writer: pd.ExcelWriter, snapshot: MetricsSnapshot) -> None:
        cycle = snapshot.cycle_metrics
        avg_s = "—"
        std_s = "—"
        if cycle.count():
            avg_s = round(cycle.average().total_seconds(), 2)
            std_s = round(cycle.std_deviation().total_seconds(), 2)

        rows = [
            ("Data", self._session_start.strftime("%Y-%m-%d %H:%M")),
            ("Duracao total (s)", round(snapshot.session_duration.total_seconds(), 1)),
            ("Ciclos completos", cycle.count()),
            ("Tempo medio ciclo (s)", avg_s),
            ("Desvio padrao ciclo (s)", std_s),
            ("% Tempo produtivo", round(snapshot.productive_percentage, 1)),
            ("% Tempo transicao", round(snapshot.transition_percentage, 1)),
            ("% Tempo interrupcao", round(snapshot.interruption_percentage, 1)),
            ("Tarefa/Zona gargalo", snapshot.bottleneck_zone or "—"),
        ]

        df = pd.DataFrame(rows, columns=["Metrica", "Valor"])
        df.to_excel(writer, sheet_name="Resumo", index=False)
        self._bold_headers(writer, "Resumo", df)

    def _write_zone_metrics(self, writer: pd.ExcelWriter, snapshot: MetricsSnapshot) -> None:
        rows = []
        for zone_name, metrics in snapshot.task_metrics.items():
            if metrics.count() == 0:
                continue
            rows.append({
                _TASK_COLUMN: zone_name,
                "Minimo (s)": round(metrics.minimum().total_seconds(), 3),
                "Medio (s)": round(metrics.average().total_seconds(), 3),
                "Maximo (s)": round(metrics.maximum().total_seconds(), 3),
                "Desvio Padrao (s)": round(metrics.std_deviation().total_seconds(), 3),
                "Ocorrencias": metrics.count(),
            })

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="Metricas por Zona", index=False)
        self._bold_headers(writer, "Metricas por Zona", df)
        self._highlight_bottleneck(writer, "Metricas por Zona", df, snapshot.bottleneck_zone)

    def _write_cycles(self, writer: pd.ExcelWriter) -> None:
        rows = []
        for cycle_result in sorted(self._cycle_results.values(), key=lambda cycle: cycle.cycle_number):
            diagnosis = _cycle_diagnosis(cycle_result)
            rows.append({
                "Nº Ciclo": cycle_result.cycle_number,
                "Inicio": cycle_result.start_time.strftime("%H:%M:%S"),
                "Fim": cycle_result.end_time.strftime("%H:%M:%S"),
                "Duracao (s)": round(cycle_result.duration.total_seconds(), 2),
                "Resultado do sistema": diagnosis.result,
                "Sequencia registada": _format_sequence(cycle_result.actual_sequence),
                "Problema detetado": diagnosis.problem,
                "Classificacao manual": "",
                "Observacoes": "",
            })

        df = pd.DataFrame(rows, columns=_CYCLE_COLUMNS)
        df.to_excel(writer, sheet_name="Ciclos", index=False)
        self._bold_headers(writer, "Ciclos", df)

    def _write_events(self, writer: pd.ExcelWriter) -> None:
        rows = [
            {
                "Ciclo": row.event.cycle_number,
                "Tarefa/Zona": row.event.zone_name,
                "Tipo": self._event_type_label(row),
                "Inicio": row.event.start_time.strftime("%H:%M:%S.%f")[:-3],
                "Fim": row.event.end_time.strftime("%H:%M:%S.%f")[:-3],
                "Duracao (s)": round(row.event.duration.total_seconds(), 3),
                "Forcado": _FORCED_LABEL[row.event.was_forced],
            }
            for row in self._events
        ]

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="Eventos", index=False)
        self._bold_headers(writer, "Eventos", df)

    def _event_type_label(self, row: _EventExportRow) -> str:
        if row.counts_as_interruption or row.event.was_forced:
            return "Interrupcao"
        return "Produtivo"

    def _bold_headers(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
        sheet = writer.sheets[sheet_name]
        for col_idx in range(1, len(df.columns) + 1):
            sheet.cell(row=1, column=col_idx).font = _HEADER_FONT

    def _highlight_bottleneck(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        df: pd.DataFrame,
        bottleneck: str | None,
    ) -> None:
        if bottleneck is None or _TASK_COLUMN not in df.columns:
            return

        sheet = writer.sheets[sheet_name]
        num_cols = len(df.columns)

        for row_idx, zone_name in enumerate(df[_TASK_COLUMN], start=2):
            if zone_name == bottleneck:
                for col_idx in range(1, num_cols + 1):
                    sheet.cell(row=row_idx, column=col_idx).fill = _BOTTLENECK_FILL
